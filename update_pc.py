# update_pc.py
import os, sys
from openpyxl import load_workbook

# workflow_dispatch で渡される inputs
row        = int(os.environ.get('INPUT_STATUS_ROW', '0'))
new_status =       os.environ.get('INPUT_NEW_STATUS', '')

if row <= 0 or not new_status:
    print("ERROR: status_row または new_status がありません。")
    sys.exit(1)

wb = load_workbook('pc_list.xlsx')
ws = wb.active
ws.cell(row=row, column=3).value = new_status   # C列をステータス列と仮定
wb.save('pc_list.xlsx')
print(f"Row {row} を '{new_status}' に更新しました。")
